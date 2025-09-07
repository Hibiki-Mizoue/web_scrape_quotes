import argparse
import csv
import os
import time
from typing import List, Dict, Optional

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BASE_URL = "https://quotes.toscrape.com"
HEADERS = {"User-Agent": "Mozilla/5.0 (portfolio-scraper; +https://github.com/yourname)"}
OUTPUT_DIR = "output"
CSV_PATH = os.path.join(OUTPUT_DIR, "quotes.csv")
XLSX_PATH = os.path.join(OUTPUT_DIR, "quotes_report.xlsx")  # 1つのファイルにシート分けで出力

# -----------------------------
# スクレイピング基本
# -----------------------------
def fetch(url: str, timeout: int = 15) -> Optional[requests.Response]:
    try:
        r = requests.get(url, headers=HEADERS, timeout=timeout)
        r.raise_for_status()
        return r
    except requests.RequestException as e:
        print(f"[WARN] fetch failed: {url} ({e})")
        return None

def parse_quotes(html: str) -> List[Dict[str, str]]:
    soup = BeautifulSoup(html, "html.parser")
    blocks = soup.select("div.quote")
    items: List[Dict[str, str]] = []
    for b in blocks:
        text = b.select_one("span.text")
        author = b.select_one("small.author")
        tags = [t.get_text(strip=True) for t in b.select("div.tags a.tag")]
        if text and author:
            items.append({
                "text": text.get_text(strip=True).strip("“”\"'"),
                "author": author.get_text(strip=True),
                "tags": ", ".join(tags)
            })
    return items

def next_page_url(html: str) -> Optional[str]:
    soup = BeautifulSoup(html, "html.parser")
    nxt = soup.select_one("li.next a")
    if nxt and nxt.get("href"):
        href = nxt["href"]
        return BASE_URL + href
    return None

def scrape_all(max_pages: Optional[int], sleep_sec: float) -> List[Dict[str, str]]:
    url = BASE_URL
    collected: List[Dict[str, str]] = []
    page = 1
    visited = set()

    while url:
        if max_pages and page > max_pages:
            break
        if url in visited:
            print("[WARN] detected loop, stop.")
            break
        visited.add(url)

        print(f"[INFO] fetch page {page}: {url}")
        res = fetch(url)
        if not res:
            print("[WARN] skip this page due to fetch error.")
            break

        rows = parse_quotes(res.text)
        if not rows:
            print("[INFO] no rows on this page, stop.")
            break
        collected.extend(rows)

        url = next_page_url(res.text)
        page += 1
        time.sleep(sleep_sec)

    # 重複除去（text + author）
    uniq = {}
    for r in collected:
        uniq[(r["text"], r["author"])] = r
    return list(uniq.values())

# -----------------------------
# 出力：CSV
# -----------------------------
def save_csv(rows: List[Dict[str, str]], path: str) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    fieldnames = ["text", "author", "tags"]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    print(f"[OK] CSV saved: {path}")

# -----------------------------
# Excel 書式（前回と同じ体裁）
#   - A1空白（B2開始）/ タイトルB2結合
#   - ヘッダー色: FFCCFFCC（薄緑）
#   - ヘッダー下は二重線 / 表全体に枠
#   - タイトルはセル幅に自動縮小で見切れ防止
# -----------------------------
def style_sheet(ws, title: str, header_fill_color="FFCCFFCC"):
    # 右端の完全空列を削除（余分列対策）
    col = ws.max_column
    while col > 1:
        empty = True
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v not in (None, ""):
                empty = False
                break
        if empty:
            ws.delete_cols(col)
        col -= 1
    max_col = ws.max_column

    # 列幅の自動調整（最低幅を確保）
    MIN_WIDTH = 12
    for c in range(2, max_col + 1):
        letter = get_column_letter(c)
        max_len = 0
        for r in range(3, ws.max_row + 1):
            v = ws.cell(row=r, column=c).value
            L = len(str(v)) if v is not None else 0
            if L > max_len:
                max_len = L
        ws.column_dimensions[letter].width = max(MIN_WIDTH, min(max_len + 2, 40))

    # タイトル（B2～最右列）を結合・配置（自動縮小で見切れ防止）
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=max_col)
    ws["B2"] = title
    ws["B2"].font = Font(size=14, bold=True)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=False, shrink_to_fit=True)
    ws.row_dimensions[2].height = 22

    thin = Side(style="thin", color="000000")
    double = Side(style="double", color="000000")

    # ヘッダー装飾（太字・背景色・下二重線・中央寄せ）
    for c in range(2, max_col + 1):
        cell = ws.cell(row=3, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=double)

    # データ枠線
    for r in range(4, ws.max_row + 1):
        for c in range(2, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# -----------------------------
# 出力：1ファイルにシート分け（名言一覧 / 著者別件数 / タグ別件数）
# -----------------------------
def save_excel_report(df_all: pd.DataFrame, out_path: str):
    # 集計データ準備
    # 著者別件数
    author_cnt = df_all.groupby("author", as_index=False).size().rename(columns={"size": "count"})
    author_cnt = author_cnt.sort_values(["count", "author"], ascending=[False, True], ignore_index=True)

    # タグ別件数（カンマ区切りを展開）
    tag_series = df_all["tags"].fillna("").astype(str)
    tag_rows = []
    for tags in tag_series:
        if tags.strip() == "":
            continue
        for t in [x.strip() for x in tags.split(",") if x.strip()]:
            tag_rows.append(t)
    if tag_rows:
        tag_cnt = pd.Series(tag_rows).value_counts().reset_index()
        tag_cnt.columns = ["tag", "count"]
    else:
        tag_cnt = pd.DataFrame(columns=["tag", "count"])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    # Writerで3シート出力（B2開始 = startrow=2, startcol=1）
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="名言一覧", index=False, startrow=2, startcol=1)
        author_cnt.to_excel(writer, sheet_name="著者別件数", index=False, startrow=2, startcol=1)
        tag_cnt.to_excel(writer, sheet_name="タグ別件数", index=False, startrow=2, startcol=1)

        ws1 = writer.sheets["名言一覧"]
        ws2 = writer.sheets["著者別件数"]
        ws3 = writer.sheets["タグ別件数"]

        # 前回と同じ色（薄緑: FFCCFFCC）で装飾
        style_sheet(ws1, "名言一覧", header_fill_color="FFCCFFCC")
        style_sheet(ws2, "著者別件数", header_fill_color="FFCCFFCC")
        style_sheet(ws3, "タグ別件数", header_fill_color="FFCCFFCC")

    print(f"[OK] Excel saved: {out_path}")

# -----------------------------
# CLI
# -----------------------------
def main():
    parser = argparse.ArgumentParser(description="Scrape quotes to CSV/Excel (single book with multiple sheets).")
    parser.add_argument("--max-pages", type=int, default=0, help="max pages to fetch (0 means all)")
    parser.add_argument("--sleep", type=float, default=0.7, help="sleep seconds per page")
    parser.add_argument("--excel", action="store_true", help="also save Excel report (one file, multiple sheets)")
    args = parser.parse_args()

    max_pages = args.max_pages if args.max_pages > 0 else None
    rows = scrape_all(max_pages, args.sleep)
    if not rows:
        print("[INFO] no data collected.")
        return

    # DataFrame 化
    df = pd.DataFrame(rows, columns=["text", "author", "tags"])

    # CSV
    save_csv(rows, CSV_PATH)

    # Excel（1ファイルにシート分け）
    if args.excel:
        try:
            save_excel_report(df, XLSX_PATH)
        except PermissionError:
            print(f"[エラー] {XLSX_PATH} が開いているため保存できません。Excel を閉じてから再実行してください。")

if __name__ == "__main__":
    main()
